"""
Excel export with deduplication.

Writes extracted form data to an Excel file, checking for duplicate
records based on Name + Roll No + Class combination.
Uses fuzzy matching on names to catch OCR variations.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from rapidfuzz import fuzz

from src.config.template_config import EXCEL_COLUMNS, FIELD_TO_EXCEL

FUZZY_THRESHOLD = 85


def _normalize(text):
    """Normalize text for comparison."""
    if not text:
        return ""
    return text.strip().upper()


def is_duplicate(existing_rows, new_record):
    """
    Check if a record already exists in the data.
    A record is duplicate if Name (fuzzy 85%+), Roll No (exact), and Class (exact) match.
    """
    new_name = _normalize(new_record.get("Name", ""))
    new_roll = _normalize(new_record.get("Roll No", ""))
    new_class = _normalize(new_record.get("Class", ""))

    if not new_name and not new_roll:
        return False

    for row in existing_rows:
        existing_name = _normalize(row.get("Name", ""))
        existing_roll = _normalize(row.get("Roll No", ""))
        existing_class = _normalize(row.get("Class", ""))

        roll_match = existing_roll == new_roll
        class_match = existing_class == new_class

        if new_name and existing_name:
            name_score = fuzz.ratio(new_name, existing_name)
            name_match = name_score >= FUZZY_THRESHOLD
        else:
            name_match = new_name == existing_name

        if name_match and roll_match and class_match:
            return True

    return False


def _build_excel_record(ocr_data):
    """Convert OCR field data to Excel column format."""
    record = {}
    for field_name, excel_col in FIELD_TO_EXCEL.items():
        if excel_col in EXCEL_COLUMNS:
            value = ocr_data.get(field_name, "")
            record[excel_col] = value.strip() if value else ""
    return record


def _load_existing_data(excel_path):
    """Load existing records from an Excel file."""
    if not os.path.exists(excel_path):
        return []

    wb = load_workbook(excel_path)
    ws = wb.active
    rows = []

    headers = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                record[headers[i]] = str(val) if val else ""
        if any(record.values()):
            rows.append(record)

    wb.close()
    return rows


def _style_worksheet(ws):
    """Apply professional styling to the worksheet."""
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    col_widths = {
        "Name": 25, "Father Name": 25, "Mother Name": 25,
        "Class": 12, "Roll No": 10, "Sec": 8,
        "Address": 35, "Date of Birth": 15,
    }
    for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = (
            col_widths.get(col_name, 15)
        )

    ws.auto_filter.ref = ws.dimensions


def export_to_excel(all_ocr_data, excel_path, on_progress=None):
    """
    Export OCR data to Excel with deduplication.

    Args:
        all_ocr_data: List of dicts, each being OCR results for one form
                      (field_name -> text).
        excel_path: Output Excel file path.
        on_progress: Optional callback(current, total, message).

    Returns:
        Tuple of (added_count, skipped_count, total_in_file).
    """
    existing_rows = _load_existing_data(excel_path)

    new_records = []
    skipped = 0
    total = len(all_ocr_data)

    for i, ocr_data in enumerate(all_ocr_data):
        record = _build_excel_record(ocr_data)

        if is_duplicate(existing_rows, record):
            skipped += 1
            if on_progress:
                on_progress(i + 1, total, f"Skipped duplicate: {record.get('Name', 'Unknown')}")
        else:
            new_records.append(record)
            existing_rows.append(record)
            if on_progress:
                on_progress(i + 1, total, f"Added: {record.get('Name', 'Unknown')}")

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Student Data"
        _style_worksheet(ws)
        start_row = 2

    data_font = Font(name="Calibri", size=11)
    data_alignment = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for row_idx, record in enumerate(new_records, start_row):
        for col_idx, col_name in enumerate(EXCEL_COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

    wb.save(excel_path)
    wb.close()

    total_in_file = len(existing_rows)
    return len(new_records), skipped, total_in_file
